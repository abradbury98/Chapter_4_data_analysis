"""
Build a Type-count heatmap comparison spreadsheet.
Sheets:
  1. All_Samples      — all 21 samples, Type × Sample count matrix
  2. ASC_CAT          — 5 ASC_CAT samples
  3. ASC_RVS          — 4 ASC_RVS samples
  4. EVI_CAT          — 6 EVI_CAT samples
  5. EVI_RVS          — 6 EVI_RVS samples
  6. CAT_comparison   — ASC_CAT vs EVI_CAT side by side
  7. RVS_comparison   — ASC_RVS vs EVI_RVS side by side

Cells are colour-gradient (white → dark blue) based on count within each sheet.
Row totals and column totals included.
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

INPUT_DIR = "/Users/alicebradbury/Desktop/ABRicate_dereplicated"
OUTPUT_PATH = os.path.join(INPUT_DIR, "TYPE_COMPARISON.xlsx")

ALL_SAMPLES = [
    "CH01_ASC_CAT", "CH02_ASC_CAT", "CH03_ASC_CAT", "CH05_ASC_CAT", "CH06_ASC_CAT",
    "CH01_ASC_RVS", "CH02_ASC_RVS", "CH05_ASC_RVS", "CH06_ASC_RVS",
    "CH01_EVI_CAT", "CH02_EVI_CAT", "CH03_EVI_CAT", "CH04_EVI_CAT", "CH05_EVI_CAT", "CH06_EVI_CAT",
    "CH01_EVI_RVS", "CH02_EVI_RVS", "CH03_EVI_RVS", "CH04_EVI_RVS", "CH05_EVI_RVS", "CH06_EVI_RVS",
]

GROUPS = {
    "ASC_CAT": ["CH01_ASC_CAT","CH02_ASC_CAT","CH03_ASC_CAT","CH05_ASC_CAT","CH06_ASC_CAT"],
    "ASC_RVS": ["CH01_ASC_RVS","CH02_ASC_RVS","CH05_ASC_RVS","CH06_ASC_RVS"],
    "EVI_CAT": ["CH01_EVI_CAT","CH02_EVI_CAT","CH03_EVI_CAT","CH04_EVI_CAT","CH05_EVI_CAT","CH06_EVI_CAT"],
    "EVI_RVS": ["CH01_EVI_RVS","CH02_EVI_RVS","CH03_EVI_RVS","CH04_EVI_RVS","CH05_EVI_RVS","CH06_EVI_RVS"],
}

# Canonical type order (grouped logically)
TYPE_ORDER = [
    "Resistance gene",
    "Efflux pump",
    "Efflux pump - Drug/Biocide",
    "Efflux pump - Drug/Metal",
    "Efflux pump - Drug/Biocide/Metal",
    "Biocide resistance",
    "Biocide/Metal resistance",
    "Metal resistance",
    "Plasmid replicon",
    "Virulence gene",
    "Virulence gene - Flagella",
    "Virulence gene - LPS/LOS",
    "Virulence gene - Surface glycosylation",
    "Virulence gene - Capsule",
    "Virulence gene - Toxin/Secretion",
    "Virulence gene - Secretion system",
    "Virulence gene - Iron acquisition",
    "Virulence gene - Invasion/Adhesion",
    "Virulence gene - Serum resistance",
    "Virulence gene - O/H antigen serotyping",
]

# Colour for each type category (left label column)
TYPE_COLOURS = {
    "Resistance gene":                        "FFF2CC",
    "Efflux pump":                            "FCE4D6",
    "Efflux pump - Drug/Biocide":             "FCE4D6",
    "Efflux pump - Drug/Metal":               "FCE4D6",
    "Efflux pump - Drug/Biocide/Metal":       "FCE4D6",
    "Biocide resistance":                     "F4CCCC",
    "Biocide/Metal resistance":               "EAD1DC",
    "Metal resistance":                       "D9D9D9",
    "Plasmid replicon":                       "CFE2F3",
    "Virulence gene":                         "D9EAD3",
    "Virulence gene - Flagella":              "B6D7A8",
    "Virulence gene - LPS/LOS":              "A2C4C9",
    "Virulence gene - Surface glycosylation": "C9DAF8",
    "Virulence gene - Capsule":               "D0E0E3",
    "Virulence gene - Toxin/Secretion":       "FF9900",
    "Virulence gene - Secretion system":      "FFD966",
    "Virulence gene - Iron acquisition":      "B45F06",
    "Virulence gene - Invasion/Adhesion":     "EA9999",
    "Virulence gene - Serum resistance":      "C27BA0",
    "Virulence gene - O/H antigen serotyping":"76A5AF",
}

# Group header colours
GROUP_HEADER_COLOURS = {
    "ASC_CAT": "2F5496",   # dark blue
    "ASC_RVS": "1F7391",   # dark teal
    "EVI_CAT": "375623",   # dark green
    "EVI_RVS": "4E6128",   # olive green
}
GROUP_TEXT = {k: "FFFFFF" for k in GROUP_HEADER_COLOURS}


def load_all_data():
    """Load all individual xlsx files, return dict sample→DataFrame."""
    data = {}
    for s in ALL_SAMPLES:
        path = os.path.join(INPUT_DIR, f"{s}.xlsx")
        if os.path.exists(path):
            df = pd.read_excel(path, sheet_name="ABRicate Results")
            data[s] = df
    return data


def build_pivot(data, samples):
    """
    Build a Type × Sample count pivot.
    Returns a DataFrame with types as index, samples as columns.
    """
    counts = {}
    for s in samples:
        if s not in data:
            counts[s] = {}
            continue
        df = data[s]
        counts[s] = df["Type"].value_counts().to_dict()

    pivot = pd.DataFrame(counts, index=TYPE_ORDER).fillna(0).astype(int)
    # Drop rows that are all zero
    pivot = pivot.loc[(pivot > 0).any(axis=1)]
    # Add totals
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot.loc["TOTAL"] = pivot.sum(axis=0)
    return pivot


def sample_group(sample_name):
    for g, members in GROUPS.items():
        if sample_name in members:
            return g
    return None


def write_heatmap_sheet(ws, pivot, title, sample_cols, data_dict=None):
    """
    Write a pivot (Type × Samples) onto worksheet ws.
    sample_cols: list of sample name columns (not TOTAL).
    Applies colour-scale conditional formatting on the data cells.
    """
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=1 + len(sample_cols) + 1)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="243F60")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Group / column headers ─────────────────────────────────────────────
    # Determine which group each sample belongs to for colour-coding header
    header_row = 2
    ws.cell(row=header_row, column=1, value="Type").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=header_row, column=1).fill = PatternFill("solid", fgColor="243F60")
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=header_row, column=1).border = border

    for ci, col in enumerate(sample_cols + ["TOTAL"], 2):
        if col == "TOTAL":
            hdr_fill = PatternFill("solid", fgColor="595959")
            hdr_font = Font(bold=True, color="FFFFFF", size=10)
        else:
            g = sample_group(col)
            hex_col = GROUP_HEADER_COLOURS.get(g, "4472C4")
            hdr_fill = PatternFill("solid", fgColor=hex_col)
            hdr_font = Font(bold=True, color="FFFFFF", size=9)

        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True, text_rotation=45)
        cell.border = border
    ws.row_dimensions[header_row].height = 80

    # ── Data rows ──────────────────────────────────────────────────────────
    all_cols = sample_cols + ["TOTAL"]
    data_start_row = header_row + 1

    for ri, (idx, row_data) in enumerate(pivot.iterrows(), data_start_row):
        is_total_row = idx == "TOTAL"

        # Type label cell
        label_cell = ws.cell(row=ri, column=1, value=idx)
        if is_total_row:
            label_cell.fill = PatternFill("solid", fgColor="595959")
            label_cell.font = Font(bold=True, color="FFFFFF")
        else:
            hex_col = TYPE_COLOURS.get(idx, "F2F2F2")
            label_cell.fill = PatternFill("solid", fgColor=hex_col)
            label_cell.font = Font(bold=not is_total_row)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = border

        for ci, col in enumerate(all_cols, 2):
            val = row_data.get(col, 0)
            cell = ws.cell(row=ri, column=ci, value=int(val) if val > 0 else "")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if is_total_row:
                cell.fill = PatternFill("solid", fgColor="595959")
                cell.font = Font(bold=True, color="FFFFFF")
            elif col == "TOTAL":
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                cell.font = Font(bold=True)

    # ── Conditional colour-scale on data cells only (not TOTAL row/col) ───
    n_types = len(pivot) - 1   # exclude TOTAL row
    n_samples = len(sample_cols)
    if n_types > 0 and n_samples > 0:
        data_range = (
            f"{get_column_letter(2)}{data_start_row}:"
            f"{get_column_letter(1 + n_samples)}{data_start_row + n_types - 1}"
        )
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="num", start_value=0,   start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="9DC3E6",
                end_type="max",                     end_color="2F5496",
            )
        )

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    for ci in range(2, 2 + len(all_cols)):
        ws.column_dimensions[get_column_letter(ci)].width = 10

    ws.freeze_panes = "B3"


def write_cross_comparison_sheet(ws, data, group_a, group_b, samples_a, samples_b):
    """
    Side-by-side comparison: samples from group_a | divider | samples from group_b.
    Each column = one sample. Colour-scale applied per group block.
    """
    pivot_a = build_pivot(data, samples_a)
    pivot_b = build_pivot(data, samples_b)

    # Union of type rows (excluding TOTAL)
    all_types = list(dict.fromkeys(
        [t for t in TYPE_ORDER if t in pivot_a.index or t in pivot_b.index]
    ))

    thin = Side(style="thin", color="CCCCCC")
    thick = Side(style="medium", color="595959")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = f"{group_a}  vs  {group_b}  —  Gene Type Comparison"
    n_a = len(samples_a)
    n_b = len(samples_b)
    total_cols = 1 + n_a + 1 + n_b  # label + groupA + divider + groupB

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(bold=True, size=13, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="243F60")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Group label row
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + n_a)
    ga_cell = ws.cell(row=2, column=2, value=group_a)
    ga_cell.fill = PatternFill("solid", fgColor=GROUP_HEADER_COLOURS[group_a])
    ga_cell.font = Font(bold=True, color="FFFFFF", size=11)
    ga_cell.alignment = Alignment(horizontal="center")

    div_col = 2 + n_a
    ws.cell(row=2, column=div_col, value="").fill = PatternFill("solid", fgColor="595959")

    ws.merge_cells(start_row=2, start_column=div_col+1, end_row=2, end_column=div_col + n_b)
    gb_cell = ws.cell(row=2, column=div_col+1, value=group_b)
    gb_cell.fill = PatternFill("solid", fgColor=GROUP_HEADER_COLOURS[group_b])
    gb_cell.font = Font(bold=True, color="FFFFFF", size=11)
    gb_cell.alignment = Alignment(horizontal="center")

    # Sample header row
    header_row = 3
    ws.cell(row=header_row, column=1, value="Type").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=header_row, column=1).fill = PatternFill("solid", fgColor="243F60")
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=header_row, column=1).border = border

    for ci, s in enumerate(samples_a, 2):
        c = ws.cell(row=header_row, column=ci, value=s)
        c.fill = PatternFill("solid", fgColor=GROUP_HEADER_COLOURS[group_a])
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, text_rotation=45)
        c.border = border

    div_c = ws.cell(row=header_row, column=div_col, value="")
    div_c.fill = PatternFill("solid", fgColor="595959")

    for ci, s in enumerate(samples_b, div_col + 1):
        c = ws.cell(row=header_row, column=ci, value=s)
        c.fill = PatternFill("solid", fgColor=GROUP_HEADER_COLOURS[group_b])
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, text_rotation=45)
        c.border = border
    ws.row_dimensions[header_row].height = 80

    # Data rows
    data_start = header_row + 1
    for ri, t in enumerate(all_types, data_start):
        lc = ws.cell(row=ri, column=1, value=t)
        hex_col = TYPE_COLOURS.get(t, "F2F2F2")
        lc.fill = PatternFill("solid", fgColor=hex_col)
        lc.font = Font(bold=True)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = border

        for ci, s in enumerate(samples_a, 2):
            val = 0
            if s in data and t in data[s]["Type"].values:
                val = int((data[s]["Type"] == t).sum())
            c = ws.cell(row=ri, column=ci, value=val if val > 0 else "")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

        ws.cell(row=ri, column=div_col, value="").fill = PatternFill("solid", fgColor="D9D9D9")

        for ci, s in enumerate(samples_b, div_col + 1):
            val = 0
            if s in data and t in data[s]["Type"].values:
                val = int((data[s]["Type"] == t).sum())
            c = ws.cell(row=ri, column=ci, value=val if val > 0 else "")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    # TOTAL row
    total_ri = data_start + len(all_types)
    ws.cell(row=total_ri, column=1, value="TOTAL").fill = PatternFill("solid", fgColor="595959")
    ws.cell(row=total_ri, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=total_ri, column=1).border = border

    for ci, s in enumerate(samples_a, 2):
        val = len(data[s]) if s in data else 0
        c = ws.cell(row=total_ri, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor="595959")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.border = border

    ws.cell(row=total_ri, column=div_col).fill = PatternFill("solid", fgColor="595959")

    for ci, s in enumerate(samples_b, div_col + 1):
        val = len(data[s]) if s in data else 0
        c = ws.cell(row=total_ri, column=ci, value=val)
        c.fill = PatternFill("solid", fgColor="595959")
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        c.border = border

    # Conditional formatting: group A block
    if len(all_types) > 0:
        range_a = (f"{get_column_letter(2)}{data_start}:"
                   f"{get_column_letter(1+n_a)}{data_start+len(all_types)-1}")
        ws.conditional_formatting.add(range_a, ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="9DC3E6",
            end_type="max",                     end_color="2F5496",
        ))
        range_b = (f"{get_column_letter(div_col+1)}{data_start}:"
                   f"{get_column_letter(div_col+n_b)}{data_start+len(all_types)-1}")
        ws.conditional_formatting.add(range_b, ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="A9D18E",
            end_type="max",                     end_color="375623",
        ))

    ws.column_dimensions["A"].width = 38
    for ci in range(2, total_cols + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = "B4"


def main():
    print("Loading sample data...")
    data = load_all_data()
    print(f"  Loaded {len(data)} samples with data.\n")

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: All samples ──────────────────────────────────────────────
    print("Building: All Samples")
    ws_all = wb.create_sheet("All_Samples")
    pivot_all = build_pivot(data, ALL_SAMPLES)
    write_heatmap_sheet(ws_all, pivot_all,
                        "All Samples — Gene Type Count Heatmap",
                        [s for s in ALL_SAMPLES if s in data])

    # ── Sheets 2-5: per group ─────────────────────────────────────────────
    for group_name, samples in GROUPS.items():
        present = [s for s in samples if s in data]
        print(f"Building: {group_name}  ({len(present)} samples)")
        ws = wb.create_sheet(group_name)
        pivot = build_pivot(data, present)
        write_heatmap_sheet(ws, pivot,
                            f"{group_name} — Gene Type Count Heatmap",
                            present)

    # ── Sheet 6: CAT comparison ───────────────────────────────────────────
    print("Building: CAT_comparison (ASC_CAT vs EVI_CAT)")
    ws_cat = wb.create_sheet("CAT_comparison")
    write_cross_comparison_sheet(
        ws_cat, data,
        "ASC_CAT", "EVI_CAT",
        GROUPS["ASC_CAT"], GROUPS["EVI_CAT"],
    )

    # ── Sheet 7: RVS comparison ───────────────────────────────────────────
    print("Building: RVS_comparison (ASC_RVS vs EVI_RVS)")
    ws_rvs = wb.create_sheet("RVS_comparison")
    write_cross_comparison_sheet(
        ws_rvs, data,
        "ASC_RVS", "EVI_RVS",
        GROUPS["ASC_RVS"], GROUPS["EVI_RVS"],
    )

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
