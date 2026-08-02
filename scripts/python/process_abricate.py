"""
Process ABRicate outputs into one dereplicated Excel spreadsheet per sample.
- Merges all databases per sample
- Deduplicates by unique gene name (keeps best %IDENTITY, then %COVERAGE)
- Adds Type column (resistance gene, efflux pump, virulence gene by subtype, etc.)
- Renames barcode IDs to sample names
"""

import pandas as pd
import os
import glob
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

BASE_DIR = "/Volumes/Alice/2_POULTRY_PROCESSING/Desktop/ABRicate_contigs"
OUTPUT_DIR = "/Users/alicebradbury/Desktop/ABRicate_dereplicated"

# ── Barcode → sample name mappings ──────────────────────────────────────────
ASC_CAT_MAP = {f"barcode{str(i).zfill(2)}": f"CH{str(i).zfill(2)}_ASC_CAT" for i in range(1, 7)}
ASC_RVS_MAP = {f"barcode{str(i).zfill(2)}": f"CH{str(i-12).zfill(2)}_ASC_RVS" for i in range(13, 19)}
EVI_CAT_MAP = {f"CH{str(i).zfill(2)}": f"CH{str(i).zfill(2)}_EVI_CAT" for i in range(1, 7)}
EVI_RVS_MAP = {f"CH{str(i).zfill(2)}": f"CH{str(i).zfill(2)}_EVI_RVS" for i in range(1, 7)}


# ── Type classification ──────────────────────────────────────────────────────
def classify_gene(row):
    db = str(row.get("DATABASE", "")).lower().strip()
    product = str(row.get("PRODUCT", "")).lower()
    gene = str(row.get("GENE", "")).lower()
    resistance = str(row.get("RESISTANCE", "")).lower()

    # ── Plasmid replicons ──
    if db == "plasmidfinder":
        return "Plasmid replicon"

    # ── O/H antigen serotyping ──
    if db == "ecoh":
        return "Virulence gene - O/H antigen serotyping"

    # ── VFDB and ecoli_vf: sub-classify by virulence category ──
    if db in ("vfdb", "ecoli_vf"):
        p = product
        # Flagella
        if ("flagell" in p or "[flagella" in p
                or any(gene.startswith(pfx) for pfx in ("fli", "flg", "fla", "mot"))
                or gene in ("chey", "chew", "chea", "rpod")):
            return "Virulence gene - Flagella"
        # LPS / LOS
        if ("[los" in p or "[lps" in p or "lipooligosaccharide" in p
                or "lipopolysaccharide" in p or "waa" in gene or "rfb" in gene):
            return "Virulence gene - LPS/LOS"
        # Pseudaminic acid / surface glycosylation
        if "pse5" in p or "glycosylat" in p or "pseudaminic" in p or "[pse5" in p:
            return "Virulence gene - Surface glycosylation"
        # Capsule
        if "[capsule" in p or "capsular" in p or "capsule biosynth" in p:
            return "Virulence gene - Capsule"
        # Toxins / cytolethal distending toxin / secreted effectors
        if ("[cdt" in p or "cytolethal" in p or "toxin" in p
                or "[ciab" in p or "[ciac" in p or "effector" in p):
            return "Virulence gene - Toxin/Secretion"
        # Secretion systems
        if "[t2ss" in p or "[t3ss" in p or "[t4ss" in p or "[t6ss" in p or "secretion system" in p:
            return "Virulence gene - Secretion system"
        # Iron acquisition / siderophores
        if ("[aerobactin" in p or "[enterobactin" in p or "iron" in p
                or "siderophore" in p or "aerobactin" in gene or "entero" in gene):
            return "Virulence gene - Iron acquisition"
        # Adhesion / invasion / fimbriae / outer membrane adhesins
        if ("[cadf" in p or "[peb1" in p or "[jlpa" in p or "[ompa" in p
                or "[momp" in p or "[agf" in p or "adhesin" in p or "invasion" in p
                or "fimbri" in p or "pilus" in p or "pilin" in p
                or "[curli" in p or "aggregative fimbri" in p):
            return "Virulence gene - Invasion/Adhesion"
        # Capsule/surface proteins not caught above
        if "[asla" in p:
            return "Virulence gene - Serum resistance"
        return "Virulence gene"

    # ── MEGARes: parse hierarchical PRODUCT field ──
    if db == "megares":
        p = product
        # Determine if efflux pump or direct resistance mechanism
        is_efflux = "efflux" in p
        is_biocide = "biocide" in p
        is_metal = "metal" in p

        if is_efflux:
            if is_biocide and is_metal:
                return "Efflux pump - Drug/Biocide/Metal"
            elif is_biocide:
                return "Efflux pump - Drug/Biocide"
            elif is_metal:
                return "Efflux pump - Drug/Metal"
            else:
                return "Efflux pump"
        if is_biocide and is_metal:
            return "Biocide/Metal resistance"
        if is_biocide:
            return "Biocide resistance"
        if is_metal:
            return "Metal resistance"
        return "Resistance gene"

    # ── CARD / NCBI / ResFinder ──
    p = product
    r = resistance

    # Efflux pumps
    if "efflux" in p:
        return "Efflux pump"

    # Metal resistance
    metal_kw = ["copper", "arsenic", "zinc", "mercury", "silver", "lead",
                "cadmium", "cobalt", "chromate", "nickel", "tellurite",
                "metal resistance"]
    if any(m in p for m in metal_kw):
        return "Metal resistance"

    # Biocide resistance
    biocide_kw = ["biocide", "quaternary ammonium", "disinfectant",
                  "chlorhexidine", "benzalkonium", "acriflavine", "qac"]
    if any(b in p or b in gene for b in biocide_kw):
        return "Biocide resistance"

    # Virulence (if labelled as such in CARD)
    if "virulence" in p or "pathogenicity" in p:
        return "Virulence gene"

    return "Resistance gene"


# ── Type → colour (hex fill) ─────────────────────────────────────────────────
TYPE_COLOURS = {
    "Resistance gene":                    "FFF2CC",   # light yellow
    "Efflux pump":                        "FCE4D6",   # light orange
    "Efflux pump - Drug/Biocide":         "FCE4D6",
    "Efflux pump - Drug/Metal":           "FCE4D6",
    "Efflux pump - Drug/Biocide/Metal":   "FCE4D6",
    "Biocide resistance":                 "F4CCCC",   # light red/pink
    "Metal resistance":                   "D9D9D9",   # light grey
    "Biocide/Metal resistance":           "EAD1DC",   # mauve
    "Plasmid replicon":                   "CFE2F3",   # light blue
    "Virulence gene":                     "D9EAD3",   # light green
    "Virulence gene - Flagella":          "B6D7A8",   # green
    "Virulence gene - LPS/LOS":           "A2C4C9",   # teal
    "Virulence gene - Surface glycosylation": "C9DAF8",
    "Virulence gene - Capsule":           "D0E0E3",
    "Virulence gene - Toxin/Secretion":   "FF9900",   # orange/amber - danger
    "Virulence gene - Secretion system":  "FFD966",
    "Virulence gene - Iron acquisition":  "B45F06",   # brown - iron
    "Virulence gene - Invasion/Adhesion": "EA9999",   # salmon
    "Virulence gene - Serum resistance":  "C27BA0",
    "Virulence gene - O/H antigen serotyping": "76A5AF",
}
DEFAULT_COLOUR = "FFFFFF"


def normalize_gene(gene):
    """Lowercase and strip trailing _N version suffixes for grouping."""
    g = str(gene).lower().strip()
    g = re.sub(r"_\d+$", "", g)
    return g


def read_abricate_tsv(path):
    """Read a single ABRicate TSV, return DataFrame or None if empty/missing."""
    if not os.path.exists(path):
        return None
    try:
        df = pd.read_csv(path, sep="\t", comment=None, low_memory=False)
        # ABRicate header starts with #FILE — rename
        df.columns = [c.lstrip("#") for c in df.columns]
        if df.empty or "FILE" not in df.columns:
            return None
        return df
    except Exception as e:
        print(f"  WARNING: could not read {path}: {e}")
        return None


def load_sample_data(tsv_paths):
    """Load and merge all TSV files for one sample into a single DataFrame."""
    frames = []
    for path in tsv_paths:
        df = read_abricate_tsv(path)
        if df is not None and not df.empty:
            frames.append(df)
    if not frames:
        return pd.DataFrame()
    merged = pd.concat(frames, ignore_index=True)
    return merged


def dereplicate(df):
    """
    Keep one row per unique (normalised) gene name.
    Best hit = highest %IDENTITY, then %COVERAGE as tiebreaker.
    """
    if df.empty:
        return df
    df = df.copy()
    df["%COVERAGE"] = pd.to_numeric(df.get("%COVERAGE", 0), errors="coerce").fillna(0)
    df["%IDENTITY"] = pd.to_numeric(df.get("%IDENTITY", 0), errors="coerce").fillna(0)
    df["_norm_gene"] = df["GENE"].apply(normalize_gene)
    # Sort so best hit is first
    df = df.sort_values(["%IDENTITY", "%COVERAGE"], ascending=False)
    df = df.drop_duplicates(subset="_norm_gene", keep="first")
    df = df.drop(columns=["_norm_gene"])
    return df


WANTED_COLS = [
    "SAMPLE", "GENE", "PRODUCT", "Type", "DATABASE",
    "%COVERAGE", "%IDENTITY", "RESISTANCE", "ACCESSION",
    "SEQUENCE", "START", "END", "STRAND",
]


def build_output_df(df, sample_name):
    """Add Sample and Type columns, select and order output columns."""
    df = df.copy()
    df["SAMPLE"] = sample_name
    df["Type"] = df.apply(classify_gene, axis=1)

    # Select only columns that exist
    cols = [c for c in WANTED_COLS if c in df.columns or c == "SAMPLE"]
    # Fill any missing wanted columns
    for c in WANTED_COLS:
        if c not in df.columns:
            df[c] = ""
    df = df[WANTED_COLS]

    # Sort by Type then Gene
    df = df.sort_values(["Type", "GENE"], key=lambda s: s.str.lower())
    return df.reset_index(drop=True)


def write_excel(df, out_path, sample_name):
    """Write DataFrame to a nicely formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ABRicate Results"

    # Header style
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        row_type = str(row.Type) if hasattr(row, "Type") else ""
        fill_hex = TYPE_COLOURS.get(row_type, DEFAULT_COLOUR)
        row_fill = PatternFill("solid", fgColor=fill_hex)

        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

    # Column widths
    col_widths = {
        "SAMPLE": 18, "GENE": 16, "PRODUCT": 55, "Type": 32,
        "DATABASE": 14, "%COVERAGE": 11, "%IDENTITY": 11,
        "RESISTANCE": 28, "ACCESSION": 16, "SEQUENCE": 14,
        "START": 10, "END": 10, "STRAND": 8,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 14)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add a legend sheet
    legend_ws = wb.create_sheet("Legend")
    legend_ws["A1"] = "Type"
    legend_ws["B1"] = "Description"
    legend_ws["A1"].font = Font(bold=True)
    legend_ws["B1"].font = Font(bold=True)
    legend_rows = [
        ("Resistance gene", "Antibiotic resistance gene (AMR)"),
        ("Efflux pump", "Drug efflux pump gene"),
        ("Efflux pump - Drug/Biocide", "Efflux pump conferring drug and biocide resistance"),
        ("Efflux pump - Drug/Metal", "Efflux pump conferring drug and metal resistance"),
        ("Efflux pump - Drug/Biocide/Metal", "Efflux pump conferring drug, biocide and metal resistance"),
        ("Biocide resistance", "Resistance to disinfectants/biocides (non-efflux)"),
        ("Metal resistance", "Heavy metal resistance"),
        ("Biocide/Metal resistance", "Combined biocide and metal resistance"),
        ("Plasmid replicon", "Plasmid incompatibility/replicon type (PlasmidFinder)"),
        ("Virulence gene", "General virulence factor"),
        ("Virulence gene - Flagella", "Flagellar motility genes"),
        ("Virulence gene - LPS/LOS", "Lipopolysaccharide / lipooligosaccharide biosynthesis"),
        ("Virulence gene - Surface glycosylation", "Surface glycoprotein modification (e.g. pseudaminic acid)"),
        ("Virulence gene - Capsule", "Capsule biosynthesis"),
        ("Virulence gene - Toxin/Secretion", "Toxin genes (e.g. CDT) and secreted effectors"),
        ("Virulence gene - Secretion system", "Type II/III/IV/VI secretion system components"),
        ("Virulence gene - Iron acquisition", "Iron uptake / siderophore genes (e.g. aerobactin, enterobactin)"),
        ("Virulence gene - Invasion/Adhesion", "Adhesins, invasins, fimbriae, outer membrane proteins"),
        ("Virulence gene - Serum resistance", "Serum/complement resistance factors"),
        ("Virulence gene - O/H antigen serotyping", "O- and H-antigen typing (EcoH database)"),
    ]
    for r_idx, (t, desc) in enumerate(legend_rows, 2):
        fill_hex = TYPE_COLOURS.get(t, DEFAULT_COLOUR)
        cell_t = legend_ws.cell(row=r_idx, column=1, value=t)
        cell_t.fill = PatternFill("solid", fgColor=fill_hex)
        cell_d = legend_ws.cell(row=r_idx, column=2, value=desc)
        cell_d.fill = PatternFill("solid", fgColor=fill_hex)
    legend_ws.column_dimensions["A"].width = 40
    legend_ws.column_dimensions["B"].width = 60

    wb.save(out_path)
    print(f"  Saved: {out_path}  ({len(df)} genes)")


# ── File discovery helpers ───────────────────────────────────────────────────

def asc_cat_files():
    """Yield (sample_name, [tsv_paths]) for ASC_CAT barcodes 01-06."""
    d = os.path.join(BASE_DIR, "ASC_CAT")
    dbs = ["card", "ecoh", "ecoli_vf", "megares", "ncbi", "plasmidfinder", "resfinder", "vfdb"]
    for bc, sample in ASC_CAT_MAP.items():
        paths = []
        for db in dbs:
            p = os.path.join(d, f"all_{bc}_porechop_nanofilt_contigs_{db}.tsv")
            if os.path.exists(p):
                paths.append(p)
        yield sample, paths


def asc_rvs_files():
    """Yield (sample_name, [tsv_paths]) for ASC_RVS barcodes 13-18 (subdirs)."""
    d = os.path.join(BASE_DIR, "ASC_RVS")
    dbs = ["card", "ncbi", "resfinder"]
    for bc, sample in ASC_RVS_MAP.items():
        paths = []
        subdir = os.path.join(d, f"all_{bc}_porechop_nanofilt")
        for db in dbs:
            p = os.path.join(subdir, f"all_{bc}_porechop_nanofilt_contigs_{db}.tsv")
            if os.path.exists(p):
                paths.append(p)
        yield sample, paths


def evi_cat_files():
    """Yield (sample_name, [tsv_paths]) for EVI_CAT CH01-CH06."""
    d = os.path.join(BASE_DIR, "EVI_CAT")
    dbs = ["card", "ncbi", "resfinder"]
    for ch_id, sample in EVI_CAT_MAP.items():
        paths = []
        for db in dbs:
            p = os.path.join(d, f"{ch_id}_EVI_CAT_contigs_{db}.tsv")
            if os.path.exists(p):
                paths.append(p)
        yield sample, paths


def evi_rvs_files():
    """Yield (sample_name, [tsv_paths]) for EVI_RVS CH01-CH06."""
    d = os.path.join(BASE_DIR, "EVI_RVS")
    dbs = ["card", "ncbi", "resfinder"]
    for ch_id, sample in EVI_RVS_MAP.items():
        paths = []
        for db in dbs:
            p = os.path.join(d, f"{ch_id}_EVI_RVS_contigs_{db}.tsv")
            if os.path.exists(p):
                paths.append(p)
        yield sample, paths


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_samples = list(asc_cat_files()) + list(asc_rvs_files()) \
                + list(evi_cat_files()) + list(evi_rvs_files())

    print(f"Processing {len(all_samples)} samples → {OUTPUT_DIR}\n")

    for sample_name, tsv_paths in all_samples:
        print(f"[{sample_name}]  ({len(tsv_paths)} database files)")
        if not tsv_paths:
            print("  No TSV files found — skipping.")
            continue

        df = load_sample_data(tsv_paths)
        if df.empty:
            print("  All files empty — skipping.")
            continue

        df = dereplicate(df)
        df = build_output_df(df, sample_name)

        out_path = os.path.join(OUTPUT_DIR, f"{sample_name}.xlsx")
        write_excel(df, out_path, sample_name)

    print("\nDone.")


if __name__ == "__main__":
    main()
