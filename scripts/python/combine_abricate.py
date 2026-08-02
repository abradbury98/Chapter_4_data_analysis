"""
Build one combined spreadsheet per sample group (ASC_CAT, ASC_RVS, EVI_CAT, EVI_RVS).
Each row = one unique gene; adds:
  - Count   : number of samples the gene was detected in
  - Samples : comma-separated list of which samples detected it
Best representative hit (highest %IDENTITY, then %COVERAGE) is kept for metadata.
"""

import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_DIR = "/Users/alicebradbury/Desktop/ABRicate_dereplicated"
OUTPUT_DIR = INPUT_DIR   # save combined files to the same folder

GROUPS = {
    "ASC_CAT": ["CH01_ASC_CAT", "CH02_ASC_CAT", "CH03_ASC_CAT",
                "CH04_ASC_CAT", "CH05_ASC_CAT", "CH06_ASC_CAT"],
    "ASC_RVS": ["CH01_ASC_RVS", "CH02_ASC_RVS", "CH03_ASC_RVS",
                "CH04_ASC_RVS", "CH05_ASC_RVS", "CH06_ASC_RVS"],
    "EVI_CAT": ["CH01_EVI_CAT", "CH02_EVI_CAT", "CH03_EVI_CAT",
                "CH04_EVI_CAT", "CH05_EVI_CAT", "CH06_EVI_CAT"],
    "EVI_RVS": ["CH01_EVI_RVS", "CH02_EVI_RVS", "CH03_EVI_RVS",
                "CH04_EVI_RVS", "CH05_EVI_RVS", "CH06_EVI_RVS"],
}

# Output column order
WANTED_COLS = [
    "GENE", "PRODUCT", "Type", "DATABASE",
    "Count", "Samples",
    "%COVERAGE", "%IDENTITY", "RESISTANCE", "ACCESSION",
]

TYPE_COLOURS = {
    "Resistance gene":                    "FFF2CC",
    "Efflux pump":                        "FCE4D6",
    "Efflux pump - Drug/Biocide":         "FCE4D6",
    "Efflux pump - Drug/Metal":           "FCE4D6",
    "Efflux pump - Drug/Biocide/Metal":   "FCE4D6",
    "Biocide resistance":                 "F4CCCC",
    "Metal resistance":                   "D9D9D9",
    "Biocide/Metal resistance":           "EAD1DC",
    "Plasmid replicon":                   "CFE2F3",
    "Virulence gene":                     "D9EAD3",
    "Virulence gene - Flagella":          "B6D7A8",
    "Virulence gene - LPS/LOS":           "A2C4C9",
    "Virulence gene - Surface glycosylation": "C9DAF8",
    "Virulence gene - Capsule":           "D0E0E3",
    "Virulence gene - Toxin/Secretion":   "FF9900",
    "Virulence gene - Secretion system":  "FFD966",
    "Virulence gene - Iron acquisition":  "B45F06",
    "Virulence gene - Invasion/Adhesion": "EA9999",
    "Virulence gene - Serum resistance":  "C27BA0",
    "Virulence gene - O/H antigen serotyping": "76A5AF",
}
DEFAULT_COLOUR = "FFFFFF"


def normalize_gene(gene):
    g = str(gene).lower().strip()
    g = re.sub(r"_\d+$", "", g)
    return g


def load_group(samples):
    """
    Load individual sample xlsx files for a group.
    Returns a DataFrame with all rows tagged with their sample name.
    Skips samples with no xlsx file (empty/no-hit samples).
    """
    frames = []
    present = []
    for s in samples:
        path = os.path.join(INPUT_DIR, f"{s}.xlsx")
        if not os.path.exists(path):
            print(f"  (no file for {s} — skipped)")
            continue
        df = pd.read_excel(path, sheet_name="ABRicate Results")
        if df.empty:
            continue
        df["_sample"] = s
        frames.append(df)
        present.append(s)
    if not frames:
        return pd.DataFrame(), present
    return pd.concat(frames, ignore_index=True), present


def build_combined(df, present_samples):
    """
    Collapse to one row per normalised gene name.
    Best hit = highest %IDENTITY then %COVERAGE across all samples.
    Adds Count and Samples columns.
    """
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["%COVERAGE"] = pd.to_numeric(df.get("%COVERAGE", 0), errors="coerce").fillna(0)
    df["%IDENTITY"] = pd.to_numeric(df.get("%IDENTITY", 0), errors="coerce").fillna(0)
    df["_norm"] = df["GENE"].apply(normalize_gene)

    # Count unique samples per normalised gene, and collect sample list
    counts = (
        df.groupby("_norm")["_sample"]
        .agg(count="nunique", samples=lambda x: ", ".join(sorted(set(x))))
        .reset_index()
    )

    # Best hit per normalised gene
    df_sorted = df.sort_values(["%IDENTITY", "%COVERAGE"], ascending=False)
    best = df_sorted.drop_duplicates(subset="_norm", keep="first").copy()

    # Merge count info onto best hits
    merged = best.merge(counts, on="_norm")
    merged = merged.rename(columns={"count": "Count", "samples": "Samples"})
    merged = merged.drop(columns=["_norm", "_sample", "SAMPLE"], errors="ignore")

    # Select and order columns
    for c in WANTED_COLS:
        if c not in merged.columns:
            merged[c] = ""
    merged = merged[WANTED_COLS]

    # Sort: Count descending (most prevalent first), then Type, then Gene
    merged = merged.sort_values(
        ["Count", "Type", "GENE"],
        ascending=[False, True, True],
        key=lambda s: s.str.lower() if s.dtype == object else s
    )
    return merged.reset_index(drop=True)


def write_excel(df, out_path, group_name, present_samples, total_samples):
    wb = Workbook()
    ws = wb.active
    ws.title = f"{group_name} Combined"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_type = str(row.Type) if hasattr(row, "Type") else ""
        fill_hex = TYPE_COLOURS.get(row_type, DEFAULT_COLOUR)
        row_fill = PatternFill("solid", fgColor=fill_hex)

        for ci, value in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

            # Bold the Count column
            if headers[ci - 1] == "Count":
                cell.font = Font(bold=True)

    # Conditional colour on Count: brighter fill for genes in all samples
    count_ci = headers.index("Count") + 1
    for ri in range(2, len(df) + 2):
        count_val = ws.cell(row=ri, column=count_ci).value
        if count_val == total_samples:
            ws.cell(row=ri, column=count_ci).fill = PatternFill("solid", fgColor="00B050")
            ws.cell(row=ri, column=count_ci).font = Font(bold=True, color="FFFFFF")
        elif isinstance(count_val, (int, float)) and count_val >= total_samples * 0.5:
            ws.cell(row=ri, column=count_ci).fill = PatternFill("solid", fgColor="92D050")

    # Column widths
    col_widths = {
        "GENE": 16, "PRODUCT": 55, "Type": 32, "DATABASE": 14,
        "Count": 8, "Samples": 55,
        "%COVERAGE": 11, "%IDENTITY": 11,
        "RESISTANCE": 28, "ACCESSION": 16,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 14)

    ws.freeze_panes = "A2"

    # Info sheet
    info_ws = wb.create_sheet("Info")
    info_ws["A1"] = f"Group: {group_name}"
    info_ws["A2"] = f"Total samples in group: {total_samples}"
    info_ws["A3"] = f"Samples with data: {len(present_samples)}"
    info_ws["A4"] = "Samples:"
    for i, s in enumerate(present_samples, 1):
        info_ws[f"B{3+i}"] = s
    info_ws["A5"] = "Count column: number of samples in which each gene was detected"
    info_ws["A6"] = "Green Count cell = detected in ALL samples; light green = ≥50% of samples"
    info_ws["A7"] = "Sorted by: Count (descending), then Type, then Gene name"
    info_ws["A1"].font = Font(bold=True, size=13)
    info_ws.column_dimensions["A"].width = 60
    info_ws.column_dimensions["B"].width = 22

    wb.save(out_path)
    print(f"  Saved: {out_path}  ({len(df)} unique genes, {len(present_samples)}/{total_samples} samples)")


def main():
    for group_name, samples in GROUPS.items():
        print(f"\n[{group_name}]")
        df_all, present = load_group(samples)
        if df_all.empty:
            print("  No data — skipping.")
            continue

        combined = build_combined(df_all, present)
        out_path = os.path.join(OUTPUT_DIR, f"COMBINED_{group_name}.xlsx")
        write_excel(combined, out_path, group_name, present, total_samples=len(samples))

    print("\nDone.")


if __name__ == "__main__":
    main()
